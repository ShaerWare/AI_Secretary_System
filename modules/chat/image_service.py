"""Chat file service: upload images/documents, OCR/text extraction, storage, cleanup."""

import hashlib
import io
import logging
import shutil
import time
from pathlib import Path
from typing import Optional

from PIL import Image


try:
    import pytesseract

    PYTESSERACT_AVAILABLE = True
except ImportError:
    pytesseract = None  # type: ignore[assignment]
    PYTESSERACT_AVAILABLE = False

try:
    import pdfplumber

    PDFPLUMBER_AVAILABLE = True
except ImportError:
    pdfplumber = None  # type: ignore[assignment]
    PDFPLUMBER_AVAILABLE = False

try:
    import openpyxl

    OPENPYXL_AVAILABLE = True
except ImportError:
    openpyxl = None  # type: ignore[assignment]
    OPENPYXL_AVAILABLE = False

try:
    import docx as python_docx

    DOCX_AVAILABLE = True
except ImportError:
    python_docx = None  # type: ignore[assignment]
    DOCX_AVAILABLE = False

logger = logging.getLogger(__name__)

IMAGES_DIR = Path("data/chat_images")
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
THUMB_MAX_WIDTH = 400

IMAGE_MIME_TYPES = {"image/jpeg", "image/png", "image/webp", "image/gif"}
DOCUMENT_MIME_TYPES = {
    "application/pdf",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # xlsx
    "application/vnd.ms-excel",  # xls
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",  # docx
    "application/msword",  # doc
    "text/plain",
    "text/csv",
    "text/markdown",
    "text/html",
    "application/json",
    "application/xml",
    "text/xml",
}
ALLOWED_MIME_TYPES = IMAGE_MIME_TYPES | DOCUMENT_MIME_TYPES

# Extension overrides for ambiguous types
_EXT_MAP = {
    "image/jpeg": "jpg",
    "image/png": "png",
    "image/webp": "webp",
    "image/gif": "gif",
    "application/pdf": "pdf",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xlsx",
    "application/vnd.ms-excel": "xls",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "docx",
    "application/msword": "doc",
    "text/plain": "txt",
    "text/csv": "csv",
    "text/markdown": "md",
    "text/html": "html",
    "application/json": "json",
    "application/xml": "xml",
    "text/xml": "xml",
}


def _generate_file_id() -> str:
    ts = str(time.time())
    return f"file_{int(time.time() * 1000)}_{hashlib.md5(ts.encode()).hexdigest()[:6]}"


def _session_dir(session_id: str) -> Path:
    """Get/create directory for session files."""
    d = IMAGES_DIR / session_id
    d.mkdir(parents=True, exist_ok=True)
    return d


def _get_extension(content_type: str, original_name: str) -> str:
    """Get file extension from MIME type or original filename."""
    if content_type in _EXT_MAP:
        return _EXT_MAP[content_type]
    # Fallback to original extension
    ext = Path(original_name).suffix.lstrip(".")
    return ext if ext else "bin"


def _is_image(content_type: str) -> bool:
    return content_type in IMAGE_MIME_TYPES


# --- Text extraction ---


def _extract_text_from_pdf(file_data: bytes) -> Optional[str]:
    """Extract text from PDF using pdfplumber."""
    if not PDFPLUMBER_AVAILABLE:
        return None
    try:
        with pdfplumber.open(io.BytesIO(file_data)) as pdf:
            pages = []
            for page in pdf.pages[:50]:  # Limit to 50 pages
                text = page.extract_text()
                if text:
                    pages.append(text)
            return "\n\n".join(pages) if pages else None
    except Exception as e:
        logger.warning(f"PDF extraction failed: {e}")
        return None


def _extract_text_from_xlsx(file_data: bytes) -> Optional[str]:
    """Extract text from Excel using openpyxl."""
    if not OPENPYXL_AVAILABLE:
        return None
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_data), read_only=True, data_only=True)
        parts = []
        for sheet_name in wb.sheetnames[:10]:  # Limit to 10 sheets
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(max_row=500, values_only=True):  # Limit to 500 rows
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    rows.append("\t".join(cells))
            if rows:
                parts.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
        wb.close()
        return "\n\n".join(parts) if parts else None
    except Exception as e:
        logger.warning(f"Excel extraction failed: {e}")
        return None


def _extract_text_from_docx(file_data: bytes) -> Optional[str]:
    """Extract text from Word document using python-docx."""
    if not DOCX_AVAILABLE:
        return None
    try:
        doc = python_docx.Document(io.BytesIO(file_data))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n\n".join(paragraphs) if paragraphs else None
    except Exception as e:
        logger.warning(f"DOCX extraction failed: {e}")
        return None


def _extract_text_from_plaintext(file_data: bytes) -> Optional[str]:
    """Extract text from plain text files (txt, csv, md, json, xml, html)."""
    try:
        text = file_data.decode("utf-8")
        return text.strip() if text.strip() else None
    except UnicodeDecodeError:
        try:
            text = file_data.decode("cp1251")
            return text.strip() if text.strip() else None
        except Exception:
            return None


def extract_document_text(file_data: bytes, content_type: str, original_name: str) -> Optional[str]:
    """Extract text from a document file based on its type."""
    ext = Path(original_name).suffix.lower()

    if content_type == "application/pdf" or ext == ".pdf":
        return _extract_text_from_pdf(file_data)

    if content_type in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ) or ext in (".xlsx", ".xls"):
        return _extract_text_from_xlsx(file_data)

    if content_type in (
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/msword",
    ) or ext in (".docx", ".doc"):
        return _extract_text_from_docx(file_data)

    # Plain text variants
    if content_type.startswith("text/") or ext in (
        ".txt",
        ".csv",
        ".md",
        ".json",
        ".xml",
        ".html",
        ".log",
        ".yaml",
        ".yml",
        ".ini",
        ".cfg",
        ".toml",
    ):
        return _extract_text_from_plaintext(file_data)

    return None


# --- Upload ---


async def upload_file(
    session_id: str,
    file_data: bytes,
    content_type: str,
    original_name: str,
) -> dict:
    """Save file, generate thumbnail (for images), extract text. Returns metadata dict."""
    if len(file_data) > MAX_FILE_SIZE:
        raise ValueError(f"File too large: {len(file_data)} > {MAX_FILE_SIZE}")

    if content_type not in ALLOWED_MIME_TYPES:
        # Try to detect by extension
        ext = Path(original_name).suffix.lower()
        text_exts = {".txt", ".csv", ".md", ".json", ".xml", ".html", ".log", ".yaml", ".yml"}
        if ext not in text_exts:
            raise ValueError(f"Unsupported type: {content_type}")

    file_id = _generate_file_id()
    ext = _get_extension(content_type, original_name)
    session_dir = _session_dir(session_id)

    # Save original
    saved_path = session_dir / f"{file_id}.{ext}"
    saved_path.write_bytes(file_data)

    is_img = _is_image(content_type)
    width = 0
    height = 0
    extracted_text: Optional[str] = None

    if is_img:
        # Image: dimensions + thumbnail + OCR
        img = Image.open(saved_path)
        width, height = img.size

        # Generate thumbnail
        thumb_path = session_dir / f"{file_id}_thumb.jpg"
        thumb = img.copy()
        if width > THUMB_MAX_WIDTH:
            ratio = THUMB_MAX_WIDTH / width
            thumb = thumb.resize((THUMB_MAX_WIDTH, int(height * ratio)), Image.LANCZOS)
        if thumb.mode in ("RGBA", "P"):
            thumb = thumb.convert("RGB")
        thumb.save(thumb_path, "JPEG", quality=80)

        # OCR
        if PYTESSERACT_AVAILABLE:
            try:
                ocr_img = img.convert("RGB") if img.mode != "RGB" else img
                raw = pytesseract.image_to_string(ocr_img, lang="rus+eng", timeout=15)
                extracted_text = raw.strip() if raw and raw.strip() else None
            except Exception as e:
                logger.warning(f"OCR failed for {original_name}: {e}")
    else:
        # Document: text extraction
        extracted_text = extract_document_text(file_data, content_type, original_name)

    return {
        "id": file_id,
        "filename": f"{file_id}.{ext}",
        "original_name": original_name,
        "size": len(file_data),
        "width": width,
        "height": height,
        "ocr_text": extracted_text,
        "mime_type": content_type,
        "is_image": is_img,
    }


# Keep backward compatibility alias
upload_image = upload_file


def get_image_path(session_id: str, filename: str) -> Optional[Path]:
    """Get full path to a file, or None if not found."""
    path = IMAGES_DIR / session_id / filename
    if path.exists() and path.is_file():
        try:
            path.resolve().relative_to(IMAGES_DIR.resolve())
            return path
        except ValueError:
            return None
    return None


def delete_session_images(session_id: str) -> int:
    """Delete all files for a session. Returns number of files deleted."""
    session_dir = IMAGES_DIR / session_id
    if not session_dir.exists():
        return 0
    count = sum(1 for f in session_dir.iterdir() if f.is_file())
    shutil.rmtree(session_dir, ignore_errors=True)
    logger.info(f"Deleted {count} files for session {session_id}")
    return count


def delete_images_by_metadata(session_id: str, metadata: dict) -> int:
    """Delete specific files referenced in message metadata."""
    images = metadata.get("images", [])
    if not images:
        return 0
    count = 0
    session_dir = IMAGES_DIR / session_id
    for img in images:
        filename = img.get("filename", "")
        image_id = img.get("id", "")
        for f in [session_dir / filename, session_dir / f"{image_id}_thumb.jpg"]:
            if f.exists():
                f.unlink(missing_ok=True)
                count += 1
    return count
