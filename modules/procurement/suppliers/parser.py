"""Config-driven readers: supplier price file -> raw rows.

Returns list of dicts with any of: article, name, price, stock, unit. The
adapter maps these to ProductOffer. Formats: `xlsx` (openpyxl), `pdf_lines`
(pdfplumber, 2-column "name .... price"). `xls`/complex-pdf added later.
"""

import logging
import re
from pathlib import Path
from typing import List, Optional

from modules.procurement.suppliers.registry import SUPPLIER_PRICES_DIR


logger = logging.getLogger(__name__)

_PRICE_TAIL_RE = re.compile(r"^(.*?)[\s ]+([\d\s .,]+)$")


def resolve_file(pattern: str) -> Optional[str]:
    """First file matching `pattern` under SUPPLIER_PRICES_DIR."""
    matches = sorted(Path(SUPPLIER_PRICES_DIR).glob(pattern))
    return str(matches[0]) if matches else None


def parse_number(raw) -> Optional[float]:
    """Parse '2 453', '1 400', '1,86', '5.3' -> float. None if not numeric."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).replace(" ", " ").strip()
    if not s:
        return None
    s = s.replace(" ", "")
    # comma as decimal separator when no dot present
    if "," in s and "." not in s:
        s = s.replace(",", ".")
    else:
        s = s.replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def _read_xlsx(path: str, cfg: dict) -> List[dict]:
    import openpyxl

    rows: List[dict] = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    cols = cfg["cols"]
    header_row = cfg.get("header_row", -1)
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_row:
            continue
        name = row[cols["name"]] if cols["name"] < len(row) else None
        article = row[cols["article"]] if cols["article"] < len(row) else None
        if not name and not article:
            continue
        price = (
            parse_number(row[cols["price"]])
            if "price" in cols and cols["price"] < len(row)
            else None
        )
        unit = row[cols["unit"]] if "unit" in cols and cols["unit"] < len(row) else None
        # optional in-file stock (some suppliers carry stock in the price file);
        # "stock2" sums a second warehouse column when present.
        stock = None
        if "stock" in cols and cols["stock"] < len(row):
            stock = parse_number(row[cols["stock"]])
            if "stock2" in cols and cols["stock2"] < len(row):
                s2 = parse_number(row[cols["stock2"]])
                if s2:
                    stock = (stock or 0) + s2
        rows.append(
            {
                "article": str(article).strip() if article else None,
                "name": str(name).strip() if name else None,
                "price": price,
                "unit": str(unit).strip() if unit else None,
                "stock": stock,
            }
        )
    wb.close()
    return rows


def _read_xlsx_stock(path: str, cfg: dict) -> dict:
    """Read the stock-only file -> {article: stock_qty}. Skips warehouse labels."""
    import openpyxl

    scols = cfg["stock_cols"]
    data_row = cfg.get("stock_data_row", 0)
    out: dict = {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < data_row:
            continue
        article = row[scols["article"]] if scols["article"] < len(row) else None
        stock = parse_number(row[scols["stock"]]) if scols["stock"] < len(row) else None
        # data rows have an article + numeric stock; warehouse labels don't
        if not article or stock is None:
            continue
        out[str(article).strip()] = stock
    wb.close()
    return out


def _read_pdf_lines(path: str, cfg: dict) -> List[dict]:
    """2-column price PDF: each product line is 'Name .... trailing_price'."""
    import pdfplumber

    rows: List[dict] = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            for line in (page.extract_text() or "").splitlines():
                line = line.strip()
                if not line:
                    continue
                m = _PRICE_TAIL_RE.match(line)
                if not m:
                    continue
                name = m.group(1).strip()
                price = parse_number(m.group(2))
                # a real product line: has a name and a plausible price
                if not name or price is None or price <= 0:
                    continue
                rows.append({"article": None, "name": name, "price": price, "unit": None})
    return rows


def _read_xls(path: str, cfg: dict) -> List[dict]:
    """Legacy .xls (1C export): category rows interspersed with products.

    A product row has a non-empty article column; category rows have a name
    but no article — tracked so each product carries its category.
    """
    import xlrd

    cols = cfg["cols"]
    header_row = cfg.get("header_row", -1)
    cat_col = cfg.get("category_col")
    require_article = cfg.get("require_article", False)
    rows: List[dict] = []
    category: Optional[str] = None
    wb = xlrd.open_workbook(path)
    ws = wb.sheet_by_index(0)
    for i in range(ws.nrows):
        if i <= header_row:
            continue

        def _cell(c: int):
            return ws.cell_value(i, c) if c < ws.ncols else None  # noqa: B023

        art = _cell(cols["article"])
        article = str(art).strip() if art not in (None, "") else None
        nm = _cell(cols["name"])
        name = str(nm).strip() if nm else None

        if cat_col is not None and name and not article:
            category = name  # category header row
            if require_article:
                continue
        if require_article and not article:
            continue
        if not name and not article:
            continue

        price = parse_number(_cell(cols["price"])) if "price" in cols else None
        rows.append(
            {"article": article, "name": name, "price": price, "category": category, "unit": None}
        )
    return rows


def _read_pdf_table(path: str, cfg: dict) -> List[dict]:
    """PDF with real column tables (pdfplumber): model/size/price + category col."""
    import pdfplumber

    cols = cfg["cols"]
    cat_col = cfg.get("category_col")
    rows: List[dict] = []
    category: Optional[str] = None
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables() or []:
                for row in table:

                    def _cell(i: int) -> str:
                        return row[i].strip() if i < len(row) and row[i] else ""  # noqa: B023

                    article = _cell(cols["article"])
                    price = parse_number(_cell(cols["price"])) if "price" in cols else None

                    if cat_col is not None and _cell(cat_col) and not article and price is None:
                        category = _cell(cat_col)  # category header row
                        continue
                    if not article or price is None or price <= 0:
                        continue

                    size = _cell(cols["size"]) if "size" in cols else ""
                    name = " ".join(p for p in (category or "", article, size) if p)
                    rows.append(
                        {
                            "article": article,
                            "name": name,
                            "price": price,
                            "category": category,
                            "unit": None,
                        }
                    )
    return rows


def parse_supplier_file(cfg: dict) -> List[dict]:
    """Parse a supplier's price file (+ optional stock file) -> raw rows."""
    path = resolve_file(cfg["price_file"])
    if not path:
        raise FileNotFoundError(f"price file not found for {cfg['key']}: {cfg['price_file']}")

    fmt = cfg["format"]
    if fmt == "xlsx":
        rows = _read_xlsx(path, cfg)
    elif fmt == "xls":
        rows = _read_xls(path, cfg)
    elif fmt == "pdf_lines":
        rows = _read_pdf_lines(path, cfg)
    elif fmt == "pdf_table":
        rows = _read_pdf_table(path, cfg)
    else:
        raise ValueError(f"unsupported format {fmt} for {cfg['key']}")

    # Merge stock by article if a stock file is configured
    if cfg.get("stock_file"):
        spath = resolve_file(cfg["stock_file"])
        if spath:
            stock_map = _read_xlsx_stock(spath, cfg)
            for r in rows:
                if r.get("article") and r["article"] in stock_map:
                    r["stock"] = stock_map[r["article"]]

    logger.info("parsed %d rows for supplier %s from %s", len(rows), cfg["key"], Path(path).name)
    return rows
